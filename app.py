import streamlit as st
import pandas as pd
import requests
import io
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
import base64
from datetime import datetime

# ── PAGE CONFIG ──────────────────────────────────────────────
st.set_page_config(
    page_title="דאשבורד קיבוצים - שאגת הארי",
    page_icon="🏘",
    layout="wide"
)

# ── RTL + CUSTOM STYLES ──────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] { direction: rtl; font-family: 'Segoe UI', Arial, sans-serif; }
h1, h2, h3 { text-align: right; }
.stMetric { direction: rtl; }
div[data-testid="metric-container"] { background: white; border-radius: 10px; padding: 14px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); }
div[data-testid="stSidebarContent"] { direction: rtl; }
.source-tag { font-size: 13px; padding: 4px 12px; border-radius: 20px; display: inline-block; margin-bottom: 10px; }
.source-live    { background: #d5f5e3; color: #1a7a44; }
.source-default { background: #f0f0f0; color: #888; }
.source-error   { background: #fdedec; color: #c0392b; }
</style>
""", unsafe_allow_html=True)


# ── FALLBACK DATA ────────────────────────────────────────────
DEFAULT_ROWS = [
    ("אילת השחר","green",86,98,184,0,0,0,0,False,7),
    ("ברעם","yellow",47,37,84,0,0,0,0,False,12),
    ("גדות","green",59,34,93,7,0,0,0,False,22),
    ("גונן","green",41,48,89,0,3,0,4,False,4),
    ("דן","yellow",84,57,141,0,0,0,4,False,15),
    ("דפנה","green",100,120,220,4,1,0,0,False,0),
    ("הגושרים","green",136,76,212,0,3,0,0,False,0),
    ("חולתה","green",117,67,184,7,0,0,0,False,120),
    ("יפתח","yellow",66,49,115,0,3,0,0,False,60),
    ("יראון","green",34,25,59,12,0,0,0,False,0),
    ("כפר בלום","green",87,93,180,0,0,0,8,False,0),
    ("כפר גלעדי","green",58,43,101,0,0,0,0,False,0),
    ("כפר הנשיא","green",62,33,95,2,0,0,0,False,0),
    ("כפר סאלד","green",70,39,109,0,0,0,0,False,0),
    ("כפר שמאי","green",60,106,166,10,0,0,0,False,0),
    ("לפידות","green",80,73,153,0,0,7,0,False,0),
    ("מנרה","green",25,24,49,0,4,0,0,False,0),
    ("מרום","red",16,8,24,0,12,0,0,False,0),
    ("מרחביה יזרעאל","green",56,27,83,0,0,4,0,False,0),
    ("מתת אל","yellow",20,14,34,0,2,0,0,False,0),
    ("נאות מרדכי","green",79,57,136,11,0,0,0,False,0),
    ("סאסא","green",37,24,61,0,0,0,0,False,0),
    ("עמיר","green",77,33,110,0,0,4,0,False,0),
    ("פלמחים","green",45,46,91,0,6,0,0,False,0),
    ("קדמת צבי","green",0,18,18,0,0,0,0,False,0),
    ("רמת נפתלי","green",95,136,231,0,5,0,0,False,0),
    ("שאר ישוב","green",70,73,143,14,0,0,0,False,0),
    ("שדה אליהו","green",53,40,93,0,0,5,0,False,0),
    ("קיבוץ נוסף","green",5,22,27,0,3,0,0,False,0),
]

COLS = ["קיבוץ","סטטוס","יסודי","נוער","סה\"כ תלמידים",
        'ש"ש',"חיילים","תנועות נוער","מכינה","סטודיו פתוח","נוער בסיכון"]

def default_df():
    df = pd.DataFrame(DEFAULT_ROWS, columns=COLS)
    df["כוח עזר סה\"כ"] = df['ש"ש'] + df["חיילים"] + df["תנועות נוער"] + df["מכינה"]
    return df


# ── ONEDRIVE UTILS ────────────────────────────────────────────
def onedrive_direct_url(share_url: str) -> str:
    encoded = base64.b64encode(share_url.encode()).decode()
    encoded = encoded.rstrip("=").replace("+", "-").replace("/", "_")
    return f"https://api.onedrive.com/v1.0/shares/u!{encoded}/root/content"


def cell_status(ws, row: int, col: int) -> str:
    try:
        fill = ws.cell(row=row, column=col).fill
        if fill.fill_type == "solid":
            rgb = fill.fgColor.rgb  # AARRGGBB
            r = int(rgb[2:4], 16)
            g = int(rgb[4:6], 16)
            if g > 180 and r < 60:  return "green"
            if r > 180 and g > 180: return "yellow"
            if r > 180 and g < 60:  return "red"
    except Exception:
        pass
    return "green"


@st.cache_data(ttl=300, show_spinner=False)
def load_from_onedrive(share_url: str) -> pd.DataFrame:
    url = onedrive_direct_url(share_url)
    resp = requests.get(url, timeout=30, allow_redirects=True)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    records = []
    for ri in range(5, 34):          # Excel rows 5–33 (29 kibbutzim)
        name = str(ws.cell(ri, 1).value or "").strip()
        if not name:
            continue
        n = lambda c: float(ws.cell(ri, c).value or 0)
        rec = {
            "קיבוץ":           name,
            "סטטוס":           cell_status(ws, ri, 7),
            "יסודי":           int(n(14)),
            "נוער":            int(n(21)),
            'סה"כ תלמידים':    int(n(22)),
            'ש"ש':             int(n(30)),
            "חיילים":          int(n(31)),
            "תנועות נוער":     int(n(32)),
            "מכינה":           int(n(33)),
            "סטודיו פתוח":     n(35) == 1,
            "נוער בסיכון":     int(n(38)),
        }
        rec['כוח עזר סה"כ'] = rec['ש"ש'] + rec["חיילים"] + rec["תנועות נוער"] + rec["מכינה"]
        records.append(rec)

    return pd.DataFrame(records) if records else None


# ── STATUS HELPERS ────────────────────────────────────────────
STATUS_LABEL = {"green": "✅ פועל תקין", "yellow": "⚠️ חלקי", "red": "🔴 לא פועל"}
STATUS_COLOR = {"green": "#27ae60", "yellow": "#f1c40f", "red": "#e74c3c"}
STATUS_BG    = {"green": "#d5f5e3", "yellow": "#fef9e7", "red": "#fdedec"}


def status_badge(s: str) -> str:
    label = STATUS_LABEL.get(s, s)
    color = STATUS_COLOR.get(s, "#999")
    bg    = STATUS_BG.get(s, "#eee")
    return f'<span style="background:{bg};color:{color};padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600">{label}</span>'


def risk_color(v: int) -> str:
    if v >= 50: return "#e74c3c"
    if v >= 10: return "#e67e22"
    if v > 0:   return "#27ae60"
    return "#bbb"


# ── SIDEBAR ───────────────────────────────────────────────────
with st.sidebar:
    st.title("⚙️ הגדרות")

    # Get URL from secrets (production) or from user input (dev/fallback)
    saved_url = st.secrets.get("ONEDRIVE_URL", "") if "ONEDRIVE_URL" in st.secrets else ""
    if saved_url:
        st.success("🔗 קישור OneDrive מוגדר")
        share_url = saved_url
    else:
        share_url = st.text_input(
            "קישור OneDrive:",
            placeholder="https://1drv.ms/x/s!...",
            help="הדבק כאן את קישור השיתוף של ניסיון.xlsx מ-OneDrive"
        )

    st.markdown("---")
    st.subheader("סינון נתונים")
    status_filter = st.multiselect(
        "סטטוס:",
        options=["green","yellow","red"],
        default=["green","yellow","red"],
        format_func=lambda x: STATUS_LABEL[x]
    )
    search_term = st.text_input("חיפוש קיבוץ:", placeholder="...")

    st.markdown("---")
    if st.button("🔄 רענן נתונים מ-OneDrive", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.markdown("---")
    st.caption("שאגת הארי | דאשבורד חינוך")


# ── LOAD DATA ─────────────────────────────────────────────────
df_raw = None
source_html = ""

if share_url:
    with st.spinner("טוען נתונים מ-OneDrive..."):
        try:
            df_raw = load_from_onedrive(share_url)
            ts = datetime.now().strftime("%d/%m/%Y %H:%M")
            source_html = f'<span class="source-tag source-live">🟢 OneDrive | עודכן: {ts}</span>'
        except Exception as e:
            st.error(f"⚠️ שגיאה בטעינה מ-OneDrive: {e}")
            source_html = '<span class="source-tag source-error">❌ שגיאה — מציג נתוני ברירת מחדל</span>'

if df_raw is None:
    df_raw = default_df()
    if not share_url:
        source_html = '<span class="source-tag source-default">📋 נתוני ברירת מחדל (הגדר קישור OneDrive)</span>'


# ── APPLY FILTERS ─────────────────────────────────────────────
df = df_raw.copy()
if status_filter:
    df = df[df["סטטוס"].isin(status_filter)]
if search_term:
    df = df[df["קיבוץ"].str.contains(search_term, na=False)]


# ── HEADER ────────────────────────────────────────────────────
st.markdown("# 🏘 דאשבורד קיבוצים — שאגת הארי")
st.markdown(source_html, unsafe_allow_html=True)
st.markdown("---")


# ── SUMMARY CARDS ─────────────────────────────────────────────
c1, c2, c3, c4, c5, c6 = st.columns(6)
green_n  = len(df_raw[df_raw["סטטוס"] == "green"])
yellow_n = len(df_raw[df_raw["סטטוס"] == "yellow"])
red_n    = len(df_raw[df_raw["סטטוס"] == "red"])
force_n  = int(df_raw['כוח עזר סה"כ'].sum())
risk_n   = int(df_raw["נוער בסיכון"].sum())

c1.metric("סה\"כ קיבוצים",      len(df_raw))
c2.metric("✅ פועלים תקין",     green_n)
c3.metric("⚠️ דורשים תשומת לב", yellow_n)
c4.metric("🔴 לא פועלים",       red_n)
c5.metric("👤 כוח עזר כולל",    force_n)
c6.metric("⚠️ נוער בסיכון",     risk_n)

st.markdown("")


# ── CHARTS ROW ────────────────────────────────────────────────
col_pie, col_bar = st.columns([1, 2])

with col_pie:
    st.subheader("התפלגות סטטוס")
    fig_pie = go.Figure(go.Pie(
        labels=["פועל תקין", "חלקי", "לא פועל"],
        values=[green_n, yellow_n, red_n],
        marker_colors=["#27ae60", "#f1c40f", "#e74c3c"],
        hole=0.55,
        textinfo="label+percent"
    ))
    fig_pie.update_layout(
        showlegend=False,
        margin=dict(t=10, b=10, l=10, r=10),
        height=250
    )
    st.plotly_chart(fig_pie, use_container_width=True)

with col_bar:
    st.subheader("נוער בסיכון לפי קיבוץ")
    risk_df = df_raw[df_raw["נוער בסיכון"] > 0].sort_values("נוער בסיכון", ascending=True)
    if len(risk_df) > 0:
        colors = [("#e74c3c" if v >= 50 else "#e67e22" if v >= 10 else "#f1c40f")
                  for v in risk_df["נוער בסיכון"]]
        fig_bar = go.Figure(go.Bar(
            x=risk_df["נוער בסיכון"],
            y=risk_df["קיבוץ"],
            orientation="h",
            marker_color=colors,
            text=risk_df["נוער בסיכון"],
            textposition="outside"
        ))
        fig_bar.update_layout(
            margin=dict(t=10, b=10, l=10, r=30),
            height=250,
            xaxis_title="",
            yaxis_title=""
        )
        st.plotly_chart(fig_bar, use_container_width=True)
    else:
        st.info("אין קיבוצים עם נוער בסיכון בסינון הנוכחי")


# ── MAIN TABLE ────────────────────────────────────────────────
st.markdown("---")
st.subheader(f"📋 נתוני קיבוצים ({len(df)} מתוך {len(df_raw)})")

if len(df) == 0:
    st.warning("לא נמצאו קיבוצים התואמים את הסינון")
else:
    # Build display dataframe
    display_rows = []
    for _, row in df.iterrows():
        # Force breakdown
        parts = []
        if row['ש"ש']:        parts.append(f'ש"ש:{row["ש"ש"]}')
        if row["חיילים"]:     parts.append(f'חיילים:{row["חיילים"]}')
        if row["תנועות נוער"]: parts.append(f'נוע׳:{row["תנועות נוער"]}')
        if row["מכינה"]:      parts.append(f'מכינה:{row["מכינה"]}')
        force_detail = " | ".join(parts) if parts else "—"

        display_rows.append({
            "קיבוץ":            row["קיבוץ"],
            "סטטוס":            STATUS_LABEL[row["סטטוס"]],
            "יסודי":            row["יסודי"] or "—",
            "נוער":             row["נוער"] or "—",
            'סה"כ תלמידים':    row['סה"כ תלמידים'] or "—",
            'כוח עזר סה"כ':    row['כוח עזר סה"כ'],
            "פירוט כוח עזר":   force_detail,
            "סטודיו פתוח":     "✅ כן" if row["סטודיו פתוח"] else "❌ לא",
            "נוער בסיכון":     row["נוער בסיכון"] if row["נוער בסיכון"] > 0 else "—",
        })

    display_df = pd.DataFrame(display_rows)

    # Color rows by status
    def style_row(row_series):
        status_val = df.iloc[row_series.name]["סטטוס"]
        bg = {"green": "#f0faf4", "yellow": "#fffde7", "red": "#fdf0ef"}.get(status_val, "")
        return [f"background-color: {bg}"] * len(row_series)

    styled = display_df.style.apply(style_row, axis=1)
    st.dataframe(styled, use_container_width=True, hide_index=True, height=600)


# ── FORCE DETAIL SECTION ─────────────────────────────────────
st.markdown("---")
with st.expander("👥 פירוט כוח עזר לפי קיבוץ"):
    force_df = df_raw[['כוח עזר סה"כ', 'ש"ש', 'חיילים', 'תנועות נוער', 'מכינה', 'קיבוץ']].copy()
    force_df = force_df[force_df['כוח עזר סה"כ'] > 0].sort_values('כוח עזר סה"כ', ascending=False)
    if len(force_df):
        fig_force = px.bar(
            force_df, x="קיבוץ",
            y=['ש"ש', 'חיילים', 'תנועות נוער', 'מכינה'],
            title="כוח עזר לפי סוג וקיבוץ",
            color_discrete_map={'ש"ש':'#3498db','חיילים':'#2ecc71','תנועות נוער':'#e67e22','מכינה':'#9b59b6'},
        )
        fig_force.update_layout(height=350, legend_title="סוג")
        st.plotly_chart(fig_force, use_container_width=True)
    else:
        st.info("אין כוח עזר בסינון הנוכחי")
